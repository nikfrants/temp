import openpyxl
import json
import os
import logging
from datetime import datetime, timedelta
import re

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Константы для названий листов и файлов
EVENTS_SHEET_NAME = 'Текущие события'
CONFIG_FILE_PATH = 'config.json'


def _parse_date_range(date_string: str) -> list:
    """
    Разбивает строку с диапазоном дат (ДД.ММ.ГГГГ) на отдельные даты.
    Например, "27.05.2025 г. - 29.05.2025 г." -> ['2025-05-27', '2025-05-28', '2025-05-29']
    """
    dates = []
    if not date_string:
        return dates

    # Заменяем потенциальные ошибки форматирования
    date_string = date_string.replace('г.', '').replace(';', '').strip()

    match = re.search(r'(\d{2}\.\d{2}\.\d{4}).*?(\d{2}\.\d{2}\.\d{4})', date_string)
    if match:
        start_date_str = match.group(1)
        end_date_str = match.group(2)

        try:
            start_date = datetime.strptime(start_date_str, '%d.%m.%Y')
            end_date = datetime.strptime(end_date_str, '%d.%m.%Y')
        except ValueError:
            logging.error(f"Некорректный формат даты в строке: '{date_string}'. Пропускаю.")
            return dates

        current_date = start_date
        while current_date <= end_date:
            dates.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)
    else:
        single_date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', date_string)
        if single_date_match:
            try:
                dates.append(datetime.strptime(single_date_match.group(1), '%d.%m.%Y').strftime('%Y-%m-%d'))
            except ValueError:
                logging.error(f"Некорректный формат даты в строке: '{date_string}'. Пропускаю.")

    return dates


def _parse_time_range(time_string: str) -> list:
    """
    Разбивает строку с временем на отдельные слоты.
    Например, "с 11:00 до 20:00" -> ["11:00-20:00"]
    """
    if not time_string:
        return []

    # Заменяем лишние символы для упрощения парсинга
    time_string = time_string.replace(';', '').strip()

    match = re.search(r'(\d{2}:\d{2})\s+до\s+(\d{2}:\d{2})', time_string)
    if match:
        start_time = match.group(1)
        end_time = match.group(2)
        return [f"{start_time}-{end_time}"]

    match_single = re.search(r'(\d{2}:\d{2})', time_string)
    if match_single:
        return [f"{match_single.group(1)}-{match_single.group(1)}"]

    return []


def _create_formatted_description(event_data: dict) -> str:
    """
    Создает форматированное описание для события на основе данных из Excel.
    """
    # Заголовки
    description = (
        f"<i>Уважаемые участники!</i>\n"
        f"<i>Ознакомьтесь с графиком и местами приема/выдачи BikeCase для соревнований “{event_data.get('name', '')}”</i>\n\n"
    )

    # Прием велосипеда в Москве
    delivery_info = []
    if event_data.get('starov_delivery'):
        delivery_info.append(f"•   <b>Староватутинский пр. 12с13</b>\n    <i>{event_data['starov_delivery']}</i>")
    if event_data.get('krylo_delivery'):
        delivery_info.append(f"•   <b>ул. Крылатская д.10</b>\n    <i>{event_data['krylo_delivery']}</i>")
    if event_data.get('starov_delivery_day_off'):
        delivery_info.append(
            f"•   <b>Староватутинский пр. 12с13 (день отъезда)</b>\n    <i>{event_data['starov_delivery_day_off']}</i>")
    if event_data.get('krylo_delivery_day_off'):
        delivery_info.append(
            f"•   <b>ул. Крылатская д.10 (день отъезда)</b>\n    <i>{event_data['krylo_delivery_day_off']}</i>")

    if delivery_info:
        description += f"📍 <i>Прием велосипеда в Москве:</i>\n{''.join(delivery_info)}\n"

    # Выдача велосипеда перед стартом
    if event_data.get('pre_start_pickup'):
        description += f"📍 <i>Выдача велосипеда перед стартом:</i>\n•   <b>{event_data.get('city', '')}</b>\n    <i>{event_data['pre_start_pickup']}</i>\n"

    # Прием велосипеда после финиша
    if event_data.get('post_finish_pickup'):
        description += f"📍 <i>Приём велосипеда после финиша:</i>\n•   <b>{event_data.get('city', '')}</b>\n    <i>{event_data['post_finish_pickup']}</i>\n"

    # Выдача велосипеда в Москве
    pickup_info = []
    if event_data.get('starov_pickup'):
        pickup_info.append(f"•   <b>Староватутинский пр. 12с13</b>\n    <i>{event_data['starov_pickup']}</i>")
    if event_data.get('krylo_pickup'):
        pickup_info.append(f"•   <b>ул. Крылатская д.10</b>\n    <i>{event_data['krylo_pickup']}</i>")

    if pickup_info:
        description += f"📍 <i>Выдача велосипеда в Москве:</i>\n{''.join(pickup_info)}"

    return description.replace('  ', ' ')  # Убираем лишние пробелы


def _create_available_slots(point_name: str, dates_str: str, times_str: str) -> dict:
    """
    Создает словарь доступных слотов с особым исключением для "Староватутинский".
    """
    if not dates_str or not times_str:
        return {}

    parsed_dates = _parse_date_range(dates_str)
    parsed_times = _parse_time_range(times_str)

    # Если даты или время не распарсились, возвращаем пустой словарь
    if not parsed_dates or not parsed_times:
        return {}

    # Специальная логика для Староватутинского - сохраняем как диапазон дат
    # Используем 'Starov' для проверки, чтобы избежать проблем с русским текстом
    if "Староватутинский" in point_name:
        date_range_string = f"{parsed_dates[0]} - {parsed_dates[-1]}"
        return {date_range_string: parsed_times}

    # Стандартная логика для всех остальных точек - один слот на каждый день
    available_slots = {}
    for date in parsed_dates:
        available_slots[date] = parsed_times

    return available_slots


def update_transfer_config(excel_file_path: str, config_file_path: str):
    """
    Обновляет JSON-конфигурацию бота, считывая данные из Excel-файла
    используя новый формат таблицы с несколькими событиями.
    """
    logging.info("Начинаю обновление конфигурации бота из Excel.")

    try:
        workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
        if EVENTS_SHEET_NAME not in workbook.sheetnames:
            logging.error(f"Лист '{EVENTS_SHEET_NAME}' не найден в файле {excel_file_path}.")
            return

        events_sheet = workbook[EVENTS_SHEET_NAME]

        # Получаем заголовки (названия событий) из первой строки
        headers = [cell.value for cell in events_sheet[1]]

        # Словарь для маппинга названий строк на ключи в JSON
        key_mapping = {
            'Внутреннее название события': 'description_raw',
            'название, вид события': 'name',
            'год договора': 'year',
            'город в которы': 'city',
            'даты/часы приема Староватутинский': 'starov_delivery',
            'даты/часы приема день отъезда СтВат': 'starov_delivery_day_off',
            'даты/часы приема Крыло': 'krylo_delivery',
            'даты/часы приема день отъезда Крыло': 'krylo_delivery_day_off',
            'даты/часы выдачи перед стартом': 'pre_start_pickup',
            'даты/часы приема после финиша': 'post_finish_pickup',
            'даты/часы выдачи Староватутинский': 'starov_pickup',
            'даты/часы выдачи Крыло': 'krylo_pickup',
        }

        new_events = []
        # Начинаем итерацию со второго столбца (индекс 1), так как первый столбец - это названия полей
        for col_idx in range(1, len(headers)):
            event_name = headers[col_idx]
            if not event_name:
                continue

            event_data_map = {}
            for row in events_sheet.iter_rows(min_row=2):
                field_name = row[0].value
                field_value = row[col_idx].value

                if field_name:
                    json_key = key_mapping.get(field_name.strip(), None)
                    if json_key:
                        event_data_map[json_key] = field_value

            if not event_data_map:
                continue

            # ИСПРАВЛЕНИЕ: Используем имя события из заголовка колонки для создания уникального ID
            event_name_for_id = event_name.lower().replace(" ", "_").replace(":", "").replace("/", "")
            event_year = event_data_map.get('year', '')
            event_id = f"{event_name_for_id}_{event_year}"

            event_json = {
                "name": event_data_map.get('name', ''),
                "id": event_id,
                # Генерируем красивое описание
                "description": _create_formatted_description(event_data_map),
                "delivery_options": [
                    {
                        "point_name": "Староватутинский пр. 12с13",
                        # Используем новую вспомогательную функцию для создания слотов
                        "available_slots": _create_available_slots("Староватутинский пр. 12с13",
                                                                   event_data_map.get('starov_delivery', ''),
                                                                   event_data_map.get('starov_delivery', ''))
                    },
                    {
                        "point_name": "ул. Крылатская д.10",
                        "available_slots": _create_available_slots("ул. Крылатская д.10",
                                                                   event_data_map.get('krylo_delivery', ''),
                                                                   event_data_map.get('krylo_delivery', ''))
                    },
                    # Добавляем новые точки, если в них есть данные
                    {
                        "point_name": "Староватутинский пр. 12с13 (день отъезда)",
                        "available_slots": _create_available_slots("Староватутинский пр. 12с13 (день отъезда)",
                                                                   event_data_map.get('starov_delivery_day_off', ''),
                                                                   event_data_map.get('starov_delivery_day_off', ''))
                    },
                    {
                        "point_name": "ул. Крылатская д.10 (день отъезда)",
                        "available_slots": _create_available_slots("ул. Крылатская д.10 (день отъезда)",
                                                                   event_data_map.get('krylo_delivery_day_off', ''),
                                                                   event_data_map.get('krylo_delivery_day_off', ''))
                    }
                ],
                "pickup_options": [
                    {
                        "point_name": "Выдача перед стартом",
                        "available_slots": _create_available_slots("Выдача перед стартом",
                                                                   event_data_map.get('pre_start_pickup', ''),
                                                                   event_data_map.get('pre_start_pickup', ''))
                    },
                    {
                        "point_name": "Приём велосипеда после финиша",
                        "available_slots": _create_available_slots("Приём велосипеда после финиша",
                                                                   event_data_map.get('post_finish_pickup', ''),
                                                                   event_data_map.get('post_finish_pickup', ''))
                    },
                    {
                        "point_name": "Староватутинский пр. 12с13",
                        "available_slots": _create_available_slots("Староватутинский пр. 12с13",
                                                                   event_data_map.get('starov_pickup', ''),
                                                                   event_data_map.get('starov_pickup', ''))
                    },
                    {
                        "point_name": "ул. Крылатская д.10",
                        "available_slots": _create_available_slots("ул. Крылатская д.10",
                                                                   event_data_map.get('krylo_pickup', ''),
                                                                   event_data_map.get('krylo_pickup', ''))
                    },
                ]
            }
            # Удаляем пустые опции
            event_json["delivery_options"] = [opt for opt in event_json["delivery_options"] if opt["available_slots"]]
            event_json["pickup_options"] = [opt for opt in event_json["pickup_options"] if opt["available_slots"]]

            new_events.append(event_json)

        new_config = {
            "admin_ids": [],
            "events": new_events
        }

        # Сохраняем или обновляем admin_ids из старого конфига
        if os.path.exists(config_file_path):
            with open(config_file_path, 'r', encoding='utf-8') as f:
                old_config = json.load(f)
                new_config["admin_ids"] = old_config.get("admin_ids", [])

        with open('config2.json', 'w', encoding='utf-8') as f:
            json.dump(new_config, f, ensure_ascii=False, indent=2)

        logging.info("Конфигурация бота успешно обновлена.")

    except FileNotFoundError:
        logging.error(f"Файл Excel '{excel_file_path}' не найден.")
    except Exception as e:
        logging.error(f"Произошла ошибка при обработке файла: {e}")
    finally:
        if 'workbook' in locals():
            workbook.close()


def main():
    """
    Главная функция для запуска скрипта из командной строки.
    """
    # Указываем путь к файлу Applications.xlsx
    excel_path = '../../database/data/applications.xlsx'

    # Имя выходного файла
    config_output_path = 'config2.json'

    update_transfer_config(excel_path, config_output_path)


if __name__ == "__main__":
    main()
