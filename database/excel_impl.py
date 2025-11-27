import os
import openpyxl
from typing import Optional
from database.abstract import Repository
from models.domain import ClientProfile, TransferApplication


class ExcelRepository(Repository):
    def __init__(self, file_path: str):
        self.file_path = file_path
        self._ensure_file_exists()

    def _ensure_file_exists(self):
        """Создает файл Excel и заголовки, если файла нет."""
        if not os.path.exists(self.file_path):
            # Создаем папку, если нет
            os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

            wb = openpyxl.Workbook()
            # Лист Клиентов
            ws_clients = wb.active
            ws_clients.title = "Clients"
            ws_clients.append(["user_id", "username", "full_name", "phone", "passport", "address"])

            # Лист Заявок
            ws_apps = wb.create_sheet("Applications")
            ws_apps.append(["id", "user_id", "event", "date", "service", "comment"])

            wb.save(self.file_path)

    async def get_user(self, user_id: int) -> Optional[ClientProfile]:
        wb = openpyxl.load_workbook(self.file_path, read_only=True)
        ws = wb["Clients"]

        found_user = None
        # Пропускаем заголовок (min_row=2)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                # Превращаем строку Excel обратно в объект
                found_user = ClientProfile(
                    user_id=row[0],
                    username=row[1],
                    full_name=row[2],
                    phone_number=str(row[3]),
                    passport_series_number=row[4],
                    registration_address=row[5]
                )
                break

        wb.close()
        return found_user

    async def save_user(self, user: ClientProfile) -> bool:
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Clients"]

        # Ищем, есть ли уже такой юзер, чтобы обновить (простая реализация - дописываем в конец пока что)
        # В идеале тут нужен поиск строки и перезапись.
        # Для старта сделаем: просто добавляем новую строку (или перезапишем, если реализуем поиск индекса)

        # Упрощенно: ищем строку
        target_row = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == user.user_id:
                target_row = idx
                break

        row_data = [
            user.user_id,
            user.username,
            user.full_name,
            user.phone_number,
            user.passport_series_number,
            user.registration_address
        ]

        if target_row:
            # Обновляем существующую
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=target_row, column=col, value=value)
        else:
            # Добавляем новую
            ws.append(row_data)

        wb.save(self.file_path)
        wb.close()
        return True

    async def create_application(self, application: TransferApplication) -> bool:
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Applications"]

        ws.append([
            application.id,
            application.user_id,
            application.event_name,
            application.dropoff_date,
            "Да" if application.tech_service_needed else "Нет",
            application.comment
        ])

        wb.save(self.file_path)
        wb.close()
        return True