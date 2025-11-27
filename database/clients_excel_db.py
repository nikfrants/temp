import openpyxl
import os
import logging
from typing import Optional, Dict, Any

logger = logging.getLogger(__name__)


class ClientsExcelManager:
    """
    Класс для управления базой клиентов в Excel-файле.
    Отвечает за чтение и запись данных клиентов.
    """

    def __init__(self, file_path: str):
        self.file_path = file_path
        # Добавляем словарь для сопоставления ключей из кода с заголовками Excel
        self._header_mapping = {
            "ID пользователя": "user_id",
            "ФИО": "full_name",
            "Номер телефона": "phone_number",
            "Серия и номер паспорта": "passport_series_number",
            "Кем выдан": "passport_issued_by",
            "Дата выдачи": "passport_date_of_issue",
            "Адрес регистрации": "registration_address",
            "История трансферов": "transfer_history",
            "Персональная скидка": "personal_discount",
            "Комментарий": "comment",
            "крыло/стар": "dropoff_point",
        }
        # Убираем загрузку рабочей книги из __init__
        # Она будет загружаться и закрываться в каждом методе

    def _setup_clients_sheet(self, workbook):
        """
        Инициализирует лист "Клиенты" с заголовками.
        """
        sheet = workbook.create_sheet("Клиенты")
        # Записываем заголовки из ключей словаря-маппинга
        headers = list(self._header_mapping.keys())
        # Записываем заголовки в первую строку
        sheet.append(headers)

    def get_user(self, user_id: int) -> Optional[Dict[str, Any]]:
        """
        Проверяет, есть ли пользователь в базе.
        Открывает, читает и закрывает файл.
        """
        if not os.path.exists(self.file_path):
            logger.warning(f"Файл {self.file_path} не найден. Пользователь не может быть найден.")
            return None

        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook["Клиенты"]
            headers = [cell.value for cell in sheet[1]]
            logger.info(f"Проверка пользователя {user_id} в Excel-БД.")

            for row in sheet.iter_rows(min_row=2):
                if row[0].value == user_id:
                    user_data = {}
                    for i, cell in enumerate(row):
                        header = headers[i]
                        python_key = self._header_mapping.get(header)
                        if python_key:
                            user_data[python_key] = cell.value
                    logger.debug(f"Пользователь ID:{user_id} найден.")
                    return user_data

            logger.debug(f"Пользователь ID:{user_id} не найден.")
            return None

        except (KeyError, FileNotFoundError) as e:
            logger.error(f"Ошибка при чтении файла {self.file_path}: {e}")
            return None
        finally:
            if 'workbook' in locals():
                workbook.close()

    def create_or_update_user(self, user_id: int, user_profile_data: dict) -> bool:
        """
        Создает нового пользователя или обновляет существующего в БД.
        Открывает, записывает и закрывает файл.
        """
        workbook = None
        try:
            if os.path.exists(self.file_path):
                workbook = openpyxl.load_workbook(self.file_path)
                sheet = workbook["Клиенты"]

                user_found = False
                for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    if row[0].value == user_id:
                        for python_key, value in user_profile_data.items():
                            excel_header = next(
                                (header for header, key in self._header_mapping.items() if key == python_key), None)
                            if excel_header:
                                try:
                                    headers = [cell.value for cell in sheet[1]]
                                    col_index = headers.index(excel_header) + 1
                                    sheet.cell(row=row_index, column=col_index, value=value)
                                except ValueError:
                                    logger.error(f"Заголовок '{excel_header}' не найден в таблице.")
                        user_found = True
                        logger.info(f"Обновлены данные пользователя ID:{user_id} в Excel-БД.")
                        break

                if not user_found:
                    headers = [cell.value for cell in sheet[1]]
                    new_row_data = [None] * len(headers)
                    for python_key, value in user_profile_data.items():
                        excel_header = next(
                            (header for header, key in self._header_mapping.items() if key == python_key), None)
                        if excel_header:
                            try:
                                col_index = headers.index(excel_header)
                                new_row_data[col_index] = value
                            except ValueError:
                                logger.error(f"Заголовок '{excel_header}' не найден в таблице.")
                    sheet.append(new_row_data)
                    logger.info(f"Создан новый пользователь ID:{user_id} в Excel-БД.")
            else:
                workbook = openpyxl.Workbook()
                self._setup_clients_sheet(workbook)
                sheet = workbook["Клиенты"]

                headers = [cell.value for cell in sheet[1]]
                new_row_data = [None] * len(headers)
                for python_key, value in user_profile_data.items():
                    excel_header = next((header for header, key in self._header_mapping.items() if key == python_key),
                                        None)
                    if excel_header:
                        try:
                            col_index = headers.index(excel_header)
                            new_row_data[col_index] = value
                        except ValueError:
                            logger.error(f"Заголовок '{excel_header}' не найден в таблице.")
                sheet.append(new_row_data)
                logger.info(f"Создан новый пользователь ID:{user_id} в Excel-БД.")

            workbook.save(self.file_path)
            return True

        except (KeyError, FileNotFoundError) as e:
            logger.error(f"Ошибка при создании/обновлении файла {self.file_path}: {e}")
            return False
        finally:
            if workbook:
                workbook.close()
